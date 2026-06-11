import io
import uuid
from datetime import date

from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import get_current_user
from app.models.user import User, UserRole
from app.services import reporting_service

router = APIRouter(prefix="/reports", tags=["reports"])


def _require_owner_or_admin(current_user: User = Depends(get_current_user)) -> User:
    if current_user.role not in (UserRole.owner, UserRole.admin, UserRole.shipper):
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Access denied")
    return current_user


@router.get("/trip/{shipment_id}")
async def trip_report(
    shipment_id: uuid.UUID,
    current_user: User = Depends(_require_owner_or_admin),
    db: AsyncSession = Depends(get_db),
):
    return await reporting_service.generate_trip_report(str(shipment_id), db)


@router.get("/fleet")
async def fleet_report(
    current_user: User = Depends(_require_owner_or_admin),
    db: AsyncSession = Depends(get_db),
):
    return await reporting_service.generate_fleet_report(str(current_user.id), db)


@router.get("/analytics")
async def platform_analytics(
    days: int = Query(30, ge=1, le=365),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != UserRole.admin:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Admin only")
    return await reporting_service.generate_platform_analytics(db, days=days)


@router.get("/advanced-analytics")
async def advanced_company_analytics(
    company_id: str = Query(..., description="Company ID for analytics"),
    days: int = Query(90, ge=7, le=365),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Advanced analytics with trends, OTIF, and corridor data for dashboard charts."""
    # Allow company members or admins to access
    if current_user.role != UserRole.admin:
        from app.models.company import CompanyMember
        member_result = await db.execute(
            select(CompanyMember).where(
                CompanyMember.company_id == company_id,
                CompanyMember.user_id == current_user.id,
                CompanyMember.is_active == True,
            )
        )
        if not member_result.scalar():
            from fastapi import HTTPException
            raise HTTPException(status_code=403, detail="Access denied")

    return await reporting_service.generate_advanced_company_analytics(company_id, db, days)


@router.get("/operational-alerts")
async def operational_alerts(
    days: int = Query(7, ge=1, le=30),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != UserRole.admin:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Admin only")
    return await reporting_service.generate_operational_alerts(db, days)


@router.get("/shipments.csv")
async def export_shipments_csv(
    current_user: User = Depends(_require_owner_or_admin),
    db: AsyncSession = Depends(get_db),
):
    owner_id = None if current_user.role == UserRole.admin else str(current_user.id)
    csv_data = await reporting_service.export_shipments_csv(owner_id, db)
    return StreamingResponse(
        io.StringIO(csv_data),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=shipments.csv"},
    )


@router.get("/shipments.pdf")
async def export_shipments_pdf(
    current_user: User = Depends(_require_owner_or_admin),
    db: AsyncSession = Depends(get_db),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from app.models.load import Load, LoadStatus
    from app.models.shipment import Shipment

    # Fetch shipments
    q = select(Shipment, Load).join(Load, Shipment.load_id == Load.id)
    if current_user.role != UserRole.admin:
        q = q.where(Load.shipper_id == current_user.id)
    q = q.order_by(Shipment.delivered_at.desc().nullslast()).limit(200)
    result = await db.execute(q)
    rows = result.all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("trakvora", styles["Title"]))
    story.append(Paragraph(f"Shipment Report — {date.today().strftime('%d %b %Y')}", styles["Normal"]))
    story.append(Paragraph(f"Account: {current_user.full_name}", styles["Normal"]))
    story.append(Spacer(1, 0.5 * cm))

    headers = ["#", "Route", "Date", "Cargo", "Weight (t)", "Amount (KES)", "Status"]
    table_data = [headers]
    for i, (shipment, load) in enumerate(rows, 1):
        delivered = shipment.delivered_at.strftime("%d/%m/%Y") if shipment.delivered_at else (load.pickup_date or "—")
        table_data.append([
            str(i),
            f"{load.pickup_location[:20]} → {load.dropoff_location[:20]}",
            delivered,
            str(load.cargo_type).replace("CargoType.", ""),
            str(load.weight_tonnes),
            f"{float(load.price_kes):,.0f}",
            str(shipment.status).replace("LoadStatus.", ""),
        ])

    col_widths = [0.8 * cm, 7 * cm, 2.2 * cm, 2.5 * cm, 2 * cm, 2.8 * cm, 2 * cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    doc.build(story)

    pdf_bytes = buf.getvalue()
    today = date.today().isoformat()
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=trakvora-shipments-{today}.pdf"},
    )


@router.get("/shipments.xlsx")
async def export_shipments_xlsx(
    current_user: User = Depends(_require_owner_or_admin),
    db: AsyncSession = Depends(get_db),
):
    """Export shipments as an Excel workbook (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from app.models.load import Load
    from app.models.shipment import Shipment

    q = select(Shipment, Load).join(Load, Shipment.load_id == Load.id)
    if current_user.role != UserRole.admin:
        q = q.where(Load.shipper_id == current_user.id)
    q = q.order_by(Shipment.delivered_at.desc().nullslast()).limit(500)
    result = await db.execute(q)
    rows = result.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shipments"

    # Header style
    header_fill = PatternFill(start_color="1a2b3c", end_color="1a2b3c", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    headers = ["#", "Shipment ID", "Route", "Cargo Type", "Weight (t)", "Amount (KES)", "Status", "Driver", "Pickup Date", "Delivered At"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for i, (shipment, load) in enumerate(rows, 2):
        delivered = shipment.delivered_at.strftime("%d/%m/%Y") if shipment.delivered_at else ""
        pickup_date = load.pickup_date.strftime("%d/%m/%Y") if load.pickup_date else ""
        route = f"{load.pickup_location} → {load.dropoff_location}"
        ws.append([
            i - 1,
            str(shipment.id),
            route,
            str(load.cargo_type).replace("CargoType.", ""),
            float(load.weight_tonnes),
            float(load.price_kes),
            str(shipment.status).replace("LoadStatus.", ""),
            "",  # driver_name placeholder (can join Driver if needed)
            pickup_date,
            delivered,
        ])

    # Column widths
    col_widths = [4, 38, 45, 15, 12, 15, 14, 20, 14, 14]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = date.today().isoformat()
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=trakvora-shipments-{today}.xlsx"},
    )


@router.get("/fleet.xlsx")
async def export_fleet_xlsx(
    current_user: User = Depends(_require_owner_or_admin),
    db: AsyncSession = Depends(get_db),
):
    """Export fleet data (trucks) as an Excel workbook."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from app.models.truck import Truck

    q = select(Truck)
    if current_user.role != UserRole.admin:
        q = q.where(Truck.owner_id == current_user.id)
    q = q.order_by(Truck.created_at.desc()).limit(500)
    result = await db.execute(q)
    trucks = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fleet"

    header_fill = PatternFill(start_color="1a2b3c", end_color="1a2b3c", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    headers = ["#", "Registration", "Type", "Make", "Model", "Year", "Capacity (t)", "Status", "Verified", "Inspection", "Insurance Expiry"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, truck in enumerate(trucks, 2):
        insurance_expiry = truck.insurance_expiry.strftime("%d/%m/%Y") if truck.insurance_expiry else ""
        ws.append([
            i - 1,
            truck.registration_number,
            str(truck.truck_type).replace("TruckType.", ""),
            truck.make or "",
            truck.model or "",
            truck.year or "",
            float(truck.capacity_tonnes) if truck.capacity_tonnes else "",
            "Active" if truck.is_active else "Inactive",
            "Yes" if truck.is_verified else "No",
            str(truck.inspection_status).replace("InspectionStatus.", ""),
            insurance_expiry,
        ])

    col_widths = [4, 18, 16, 14, 14, 8, 14, 10, 10, 18, 16]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = date.today().isoformat()
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=trakvora-fleet-{today}.xlsx"},
    )


@router.get("/shipments/{shipment_id}/invoice.pdf")
async def export_invoice_pdf(
    shipment_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from app.models.load import Load
    from app.models.shipment import Shipment

    result = await db.execute(
        select(Shipment, Load).join(Load, Shipment.load_id == Load.id).where(Shipment.id == shipment_id)
    )
    row = result.first()
    if not row:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Shipment not found")
    shipment, load = row

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("trakvora", styles["Title"]))
    story.append(Paragraph("Invoice / Shipment Summary", styles["Heading2"]))
    story.append(Spacer(1, 0.4 * cm))

    # Look up VAT rate for the shipper's country
    from app.models.country_config import CountryConfig
    from app.models.user import User as UserModel
    shipper_result = await db.execute(select(UserModel).where(UserModel.id == load.shipper_id))
    shipper = shipper_result.scalar_one_or_none()
    shipper_country = (shipper.country if shipper else None) or "KE"
    cc = (await db.execute(
        select(CountryConfig).where(CountryConfig.country_code == shipper_country.upper())
    )).scalar_one_or_none()
    vat_rate = float(cc.vat_rate) if cc else 0.16
    price = float(load.price_kes)
    vat_amount = round(price * vat_rate, 2)
    total_with_vat = round(price + vat_amount, 2)

    delivered = shipment.delivered_at.strftime("%d %b %Y") if shipment.delivered_at else "In progress"
    detail_rows = [
        ["Shipment ID", str(shipment.id)],
        ["Status", str(shipment.status).replace("LoadStatus.", "").title()],
        ["From", load.pickup_location],
        ["To", load.dropoff_location],
        ["Cargo", f"{load.cargo_type} — {load.weight_tonnes}t"],
        ["Pickup Date", load.pickup_date or "—"],
        ["Delivered", delivered],
        ["Subtotal (KES)", f"{price:,.2f}"],
        [f"VAT ({vat_rate * 100:.0f}%)", f"{vat_amount:,.2f}"],
        ["Total incl. VAT (KES)", f"{total_with_vat:,.2f}"],
    ]
    t = Table(detail_rows, colWidths=[4 * cm, 13 * cm])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph("Generated by trakvora — www.trakvora.com", styles["Normal"]))

    doc.build(story)
    pdf_bytes = buf.getvalue()
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=invoice-{shipment_id}.pdf"},
    )


@router.get("/commissions")
async def commission_reconciliation(
    period: str = Query(None, description="YYYY-MM — filters to that calendar month"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Commission reconciliation report for admins.
    Returns totals + per-partner breakdown for a given month (or all-time if no period).
    """
    if current_user.role != UserRole.admin:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Admin access required")

    from datetime import datetime, timezone
    from sqlalchemy import func, case
    from app.models.commission_invoice import CommissionInvoice, CommissionInvoiceStatus
    from app.models.user import User as UserModel

    # Build date range filter
    date_filter = []
    if period:
        try:
            year, month = int(period[:4]), int(period[5:7])
            from calendar import monthrange
            last_day = monthrange(year, month)[1]
            start = datetime(year, month, 1, tzinfo=timezone.utc)
            end = datetime(year, month, last_day, 23, 59, 59, tzinfo=timezone.utc)
            date_filter = [CommissionInvoice.due_at >= start, CommissionInvoice.due_at <= end]
        except (ValueError, IndexError):
            pass

    # Aggregate totals
    total_q = select(
        func.coalesce(func.sum(CommissionInvoice.amount_kes), 0).label("total_invoiced"),
        func.coalesce(
            func.sum(case((CommissionInvoice.status == CommissionInvoiceStatus.paid, CommissionInvoice.amount_kes), else_=0)), 0
        ).label("total_collected"),
        func.coalesce(
            func.sum(case((CommissionInvoice.status == CommissionInvoiceStatus.overdue, CommissionInvoice.amount_kes), else_=0)), 0
        ).label("total_overdue"),
    ).where(*date_filter)
    totals = (await db.execute(total_q)).one()

    total_invoiced  = float(totals.total_invoiced)
    total_collected = float(totals.total_collected)
    total_overdue   = float(totals.total_overdue)
    total_outstanding = total_invoiced - total_collected

    # Per-partner breakdown
    partner_q = (
        select(
            CommissionInvoice.owner_id,
            func.count().label("invoice_count"),
            func.coalesce(func.sum(CommissionInvoice.amount_kes), 0).label("invoiced"),
            func.coalesce(
                func.sum(case((CommissionInvoice.status == CommissionInvoiceStatus.paid, CommissionInvoice.amount_kes), else_=0)), 0
            ).label("paid"),
            func.coalesce(
                func.sum(case((CommissionInvoice.status == CommissionInvoiceStatus.overdue, CommissionInvoice.amount_kes), else_=0)), 0
            ).label("overdue"),
            func.bool_or(CommissionInvoice.auto_blocked).label("is_blocked"),
        )
        .where(*date_filter)
        .group_by(CommissionInvoice.owner_id)
        .order_by(func.sum(CommissionInvoice.amount_kes).desc())
    )
    partner_rows = (await db.execute(partner_q)).all()

    # Load owner names in bulk
    owner_ids = [r.owner_id for r in partner_rows]
    owners = {}
    if owner_ids:
        owner_result = (await db.execute(
            select(UserModel.id, UserModel.full_name).where(UserModel.id.in_(owner_ids))
        )).all()
        owners = {str(r.id): r.full_name for r in owner_result}

    by_partner = []
    for r in partner_rows:
        inv = float(r.invoiced)
        pd  = float(r.paid)
        ov  = float(r.overdue)
        outstanding = inv - pd
        status = "clear" if outstanding <= 0 else ("overdue" if ov > 0 else "partial")
        if r.is_blocked:
            status = "blocked"
        by_partner.append({
            "owner_id":       str(r.owner_id),
            "owner_name":     owners.get(str(r.owner_id), "Unknown"),
            "invoice_count":  r.invoice_count,
            "invoiced_kes":   round(inv, 2),
            "paid_kes":       round(pd, 2),
            "outstanding_kes": round(outstanding, 2),
            "overdue_kes":    round(ov, 2),
            "status":         status,
        })

    return {
        "period":               period or "all-time",
        "total_invoiced_kes":   round(total_invoiced, 2),
        "total_collected_kes":  round(total_collected, 2),
        "total_outstanding_kes": round(total_outstanding, 2),
        "total_overdue_kes":    round(total_overdue, 2),
        "by_partner":           by_partner,
    }


@router.get("/commissions.xlsx")
async def commission_reconciliation_xlsx(
    period: str = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Download commission reconciliation as Excel."""
    if current_user.role != UserRole.admin:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Admin access required")

    # Reuse the JSON endpoint logic
    from fastapi import Request
    data = await commission_reconciliation(period=period, current_user=current_user, db=db)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=503, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commission Reconciliation"

    # Summary section
    summary_header = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    ws.append(["Commission Reconciliation", data["period"]])
    ws.append(["Total Invoiced (KES)", data["total_invoiced_kes"]])
    ws.append(["Total Collected (KES)", data["total_collected_kes"]])
    ws.append(["Total Outstanding (KES)", data["total_outstanding_kes"]])
    ws.append(["Total Overdue (KES)", data["total_overdue_kes"]])
    ws.append([])

    # Per-partner table header
    headers = ["Partner", "Invoices", "Invoiced (KES)", "Paid (KES)", "Outstanding (KES)", "Overdue (KES)", "Status"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = summary_header
        cell.fill = header_fill

    for row in data["by_partner"]:
        ws.append([
            row["owner_name"],
            row["invoice_count"],
            row["invoiced_kes"],
            row["paid_kes"],
            row["outstanding_kes"],
            row["overdue_kes"],
            row["status"],
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"commissions-{period or 'all-time'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/trips/{shipment_id}/dwell")
async def dwell_report(
    shipment_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Dwell time analysis for a shipment — stops detected from GPS trail.
    Accessible by the shipment's driver, owner, shipper, or admin.
    """
    from app.models.shipment import Shipment as ShipmentModel
    from app.models.load import Load as LoadModel

    shipment = (await db.execute(
        select(ShipmentModel).where(ShipmentModel.id == shipment_id)
    )).scalar_one_or_none()
    if not shipment:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Shipment not found")

    if current_user.role != UserRole.admin:
        allowed = {shipment.driver_id, shipment.owner_id}
        load = (await db.execute(
            select(LoadModel).where(LoadModel.id == shipment.load_id)
        )).scalar_one_or_none()
        if load:
            allowed.add(load.shipper_id)
        if current_user.id not in allowed:
            from fastapi import HTTPException
            raise HTTPException(status_code=403, detail="Access denied")

    return await reporting_service.generate_dwell_report(str(shipment_id), db)
